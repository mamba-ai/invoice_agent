from PIL import Image
import ast
import os 
import io 
import base64
import imghdr
from openai import OpenAI

import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink

import streamlit as st 
from surya.ocr import run_ocr
from surya.model.detection import model
from surya.model.recognition.model import load_model as load_rec_model
from surya.model.recognition.processor import load_processor as load_rec_processor
import openai
import anthropic


OPENAI_API_KEY = os.environ.get('OPENAI_API_KEY', None)
client = openai.OpenAI(api_key=OPENAI_API_KEY)
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", None)
claude_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
OPENROUTER_API_KEY = os.environ.get("OPENROUTER_API_KEY", None)
openrouter_client = OpenAI(
  base_url="https://openrouter.ai/api/v1",
  api_key=OPENROUTER_API_KEY,
)
LANGUAGES = ["ja", "en"]
json_example = '''
{
     "title": "領収書/請求書",
     "recipient": "OOXX株式会社",
     "issue_date": "OO年OO月OO日",
     "description": "飲食代",
     "issuer": "XXOO株式会社",
     "post_code": "123-4567",
     "address": "東京都千代田区駿河台2-2",
     "tel": "987-6543-210",
     "registration_number": "T9010901044466", # optonal
     "before_tax_10_percentage": 60000, # optonal
     "tax_10_percentage": 6000, # optonal
     "before_tax_8_percentage": 10000, # optonal
     "tax_8_percentage": 800, # optonal
     "amount_before_tax": 70000, # optonal
     "total_amount": 76800,
     "items": [],
}
'''


def pil_image_to_base64(pil_image):
    """
    Convert a PIL image to a base64 string.

    :param pil_image: PIL image object.
    :return: Base64 encoded string of the image.
    """
    # Create a BytesIO object to store the image data
    image_data = io.BytesIO()
    # Save the image to the BytesIO object in PNG format
    pil_image.save(image_data, format="PNG")
    # Convert the image data to a base64 string
    base64_encoded = base64.b64encode(image_data.getvalue())
    # Convert the base64 bytes to a string
    base64_string = base64_encoded.decode('utf-8')
    return base64_string


def detect_image_type(base64_str):
    # Decode the base64 string to binary data
    image_data = base64.b64decode(base64_str)
    # Detect the image type
    image_type = imghdr.what(None, h=image_data)
    return image_type


def ocr_invoice(base64_image_string):
    """
    Perform OCR on an invoice image and convert the result to a JSON object.

    :param image_path: Path to the invoice image file.
    :return: JSON object containing the parsed invoice data.
    """
    # Convert the image to a base64 string
    image_data = base64_image_string
    image_type = detect_image_type(image_data)
    media_type = f"image/{image_type}"
    
    # Perform OCR on the image
    message = claude_client.messages.create(
        model="claude-3-5-sonnet-20240620",
        max_tokens=3000,
        temperature=0,
        system="You are a helpful OCR assistant designed to recognize content from image and output JSON.",
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type": "base64",
                            "media_type": media_type,
                            "data": image_data,
                        },
                    },
                    {
                        "type": "text",
                        "text": '''You are OCR expert, parse, detect, recognize and convert following receipt/invoice image result into structure receipt data object. 
    Don't make up value not in the Input. Output must be a well-formed JSON object. And both key and values in the json object should be in Japanese or Number. Don't lose any information. 
    Only response json object is needed. No need to response any other information.'''
                    }
                ]
            }
        ]
    )
    # Return the content of the message
    return message.content[0].text

def ocr_invoice_openrouter(base64_image_string):
    """
    Perform OCR on an invoice image and convert the result to a JSON object.

    :param image_path: Path to the invoice image file.
    :return: JSON object containing the parsed invoice data.
    """
    # Convert the image to a base64 string
    # base64_image_string = base64_to_webp_and_back(base64_image_string)
    image_type = detect_image_type(base64_image_string)
    media_type = f"image/{image_type}"
    
    completion = openrouter_client.chat.completions.create(
        extra_headers={
            "X-Title": "MAMBA-AI", # Optional. Shows in rankings on openrouter.ai.
        },
        model="anthropic/claude-3.5-sonnet",
        max_tokens=3000,
        temperature=0,
        response_format={ "type": "json_object" },
        messages=[
            {
                "role": "system",
                "content": "You are a helpful OCR assistant designed to recognize content from image and output JSON.",
            },
            {
                "role": "user",
                "content":[
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:{media_type};base64,{base64_image_string}",
                        }
                    },
                    {
                        "type": "text",
                        "text": f'''You are OCR expert, parse, detect, recognize and convert following receipt/invoice image result into structure receipt data object.
Don't make up value not in the Input. Don't lose any information.
Only response json object is needed. No need to response any other information.
Output must be a well-formed JSON format. Following is an example of the JSON object: 
{json_example}
If you can't find the value, please leave it null.
'''
                    }
                ]
            }
        ],
    )
    return completion.choices[0].message.content


def get_json_result_v2(pil_image, models):
    base64_image_string = pil_image_to_base64(pil_image)
    # json_result = ocr_invoice_openrouter(base64_image_string)
    json_result = ocr_invoice(base64_image_string)
    return json_result 

@st.cache_resource()
def load_models():
    det_model, det_processor = model.load_model(), model.load_processor()
    rec_model, rec_processor = load_rec_model(), load_rec_processor()
    return det_processor, det_model, rec_model, rec_processor

def get_ocr_predictions(pil_image, models):
    det_processor, det_model, rec_model, rec_processor = models
    img_pred = run_ocr([pil_image], [LANGUAGES], det_model, det_processor, rec_model, rec_processor)[0]
    predictions = []
    for line in img_pred.text_lines:
        predictions.append([ast.literal_eval(str(line.bbox)), str(line.text)])
    print(predictions)
    return predictions

def get_json_result(predictions):
    user_prompt = f"""
    ### Instruction:
    You are POS receipt data expert, parse, detect, recognize and convert following receipt OCR image result into structure receipt data object. 
    Don't make up value not in the Input. Output must be a well-formed JSON object. And both key and values in the json object should be in Japanese or Number. Don't lose any information. 
    Do not nest JSON objects more than two levels deep and Follow the same structure as the input.
    ```json

    ### Input:
    {predictions}
    
    ### Output:
    """
    response = client.chat.completions.create(
        model="gpt-4o",
        temperature=0.0,
        response_format={ "type": "json_object" },
        messages=[
            {"role": "system", "content": "You are a helpful assistant designed to output JSON."},
            {"role": "user", "content": user_prompt}
        ]
    )
    json_result = response.choices[0].message.content
    return json_result
    
    
def json_to_excel_with_links_v2(json_data, excel_file_path):
    """
    Convert JSON data to Excel file with nested JSONs in new sheets and hyperlinks.
    
    Parameters:
    json_data (str or dict): JSON data as a string or dictionary.
    excel_file_path (str): Path where the Excel file will be saved.
    """
    if isinstance(json_data, str):
        json_data = json.loads(json_data)
    
    wb = Workbook()
    main_sheet = wb.active
    main_sheet.title = 'Main'

    def add_sheet(sheet_name, data):
        """
        Add a new sheet with given data.
        
        Parameters:
        sheet_name (str): The name of the new sheet.
        data (list or dict): The data to be added to the new sheet.
        """
        df = pd.DataFrame(data)
        # if isinstance(data, list):
            # df.insert(0, 'Sequence', range(1, len(df) + 1))
        
        new_sheet = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)

    def normalize_json(json_obj, sheet_name_prefix=''):
        """
        Normalize JSON object to handle nested structures and add new sheets.
        
        Parameters:
        json_obj (dict or list): JSON object or list of JSON objects.
        sheet_name_prefix (str): Prefix for naming nested sheets.
        
        Returns:
        dict: Normalized JSON object.
        """
        if isinstance(json_obj, dict):
            normalized_dict = {}
            for k, v in json_obj.items():
                if isinstance(v, dict):
                    sheet_name = f"{sheet_name_prefix}{k}"
                    add_sheet(sheet_name, [v])
                    normalized_dict[k] = f'[{sheet_name}]'
                elif isinstance(v, list) and all(isinstance(i, dict) for i in v):
                    sheet_name = f"{sheet_name_prefix}{k}"
                    add_sheet(sheet_name, v)
                    normalized_dict[k] = f'[{sheet_name}]'
                else:
                    normalized_dict[k] = v
            return normalized_dict
        elif isinstance(json_obj, list):
            return [normalize_json(item, sheet_name_prefix) for item in json_obj]
        else:
            return json_obj

    normalized_data = []
    if isinstance(json_data, list):
        normalized_data = normalize_json(json_data)
    else:
        normalized_data.append(normalize_json(json_data))
    
    df_main = pd.DataFrame(normalized_data)
    for r in dataframe_to_rows(df_main, index=False, header=True):
        main_sheet.append(r)
    
    # Add hyperlinks to the main sheet
    for row in main_sheet.iter_rows(min_row=2, max_col=len(df_main.columns), max_row=main_sheet.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('[') and cell.value.endswith(']'):
                sheet_name = cell.value.strip('[]')
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.value = sheet_name

    wb.save(excel_file_path)
    
    
def json_to_excel_with_links(json_data, excel_file_path):
    """
    Convert JSON data to Excel file with nested JSONs in new sheets and hyperlinks.
    
    Parameters:
    json_data (str or dict): JSON data as a string or dictionary.
    excel_file_path (str): Path where the Excel file will be saved.
    """
    if isinstance(json_data, str):
        json_data = json.loads(json_data)
    
    wb = Workbook()
    main_sheet = wb.active
    main_sheet.title = 'Main'
    
    def add_sheet(sheet_name, data):
        """
        Add a new sheet with given data.
        
        Parameters:
        sheet_name (str): The name of the new sheet.
        data (list or dict): The data to be added to the new sheet.
        """
        df = pd.DataFrame(data)
        new_sheet = wb.create_sheet(title=sheet_name)
        for r in dataframe_to_rows(df, index=False, header=True):
            new_sheet.append(r)

    def process_nested_json(json_obj, parent_sheet_name):
        """
        Process nested JSON and add data to new sheets.
        
        Parameters:
        json_obj (dict or list): JSON object or list of JSON objects.
        parent_sheet_name (str): Parent sheet name for creating new sheet names.
        
        Returns:
        dict: Normalized JSON object with hyperlinks for nested structures.
        """
        if isinstance(json_obj, dict):
            normalized_dict = {}
            for k, v in json_obj.items():
                if isinstance(v, (dict, list)):
                    sheet_name = f"{k}"
                    if isinstance(v, dict):
                        add_sheet(sheet_name, [v])
                    else:
                        add_sheet(sheet_name, v)
                    normalized_dict[k] = f'[{sheet_name}]'
                    process_nested_json(v, sheet_name)
                else:
                    normalized_dict[k] = v
            return normalized_dict
        elif isinstance(json_obj, list):
            return [process_nested_json(item, parent_sheet_name) for item in json_obj]
        else:
            return json_obj
    
    normalized_data = []
    if isinstance(json_data, list):
        for idx, item in enumerate(json_data):
            normalized_data.append(process_nested_json(item, f"Main_{idx}"))
    else:
        normalized_data.append(process_nested_json(json_data, "Main"))
    
    df_main = pd.DataFrame(normalized_data)
    for r in dataframe_to_rows(df_main, index=False, header=True):
        main_sheet.append(r)
    
    # Add hyperlinks to the main sheet
    for row in main_sheet.iter_rows(min_row=2, max_col=len(df_main.columns), max_row=main_sheet.max_row):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith('[') and cell.value.endswith(']'):
                sheet_name = cell.value.strip('[]')
                cell.hyperlink = f"#{sheet_name}!A1"
                cell.value = sheet_name

    wb.save(excel_file_path)
    
    
if __name__ == "__main__":
    json_data = '''
    [
        {"name": "John", "age": 30, "city": "New York"},
        {"name": "Anna", "age": 22, "city": "London"},
        {"name": "Mike", "age": 32, "city": "San Francisco"}
    ]
    '''
    json_data_2 = '''
        {
  "請求書": "御請求書",
  "会社情報": {
    "会社名": "株式会社SNSソフト",
    "郵便番号": "〒101-0031",
    "住所": "東京都千代田区東神田・TQ東神田ビル2階",
    "電話番号": "03-1234-5678"
  },
  "宛先": "株式会社ABC 御中",
  "挨拶文": "平素は格別のご高配に賜り、誠にありがとう御座います。",
  "担当": "担当",
  "請求内容": "下記の通りご請求申し上げます。",
  "合計金額": "¥550,000（消費税含）",
  "お支払期日": "2024年5月31日",
  "明細": [
    {
      "商品番号": "1",
      "商品名": "システム開発支援",
      "作業年月": "2024年4月",
      "数量": "1.0",
      "商品単価": "500,000",
      "金額": "500,000",
      "備考": ""
    },
    {
      "商品番号": "2",
      "商品名": "システム開発支援",
      "作業年月": "2024年4月",
      "数量": "1.0",
      "商品単価": "500,000",
      "金額": "500,000",
      "備考": ""
    }
  ],
  "小計": "500,000",
  "消費税": "50,000",
  "合計金額再掲": "550,000",
  "振込先": {
    "銀行名": "○○銀行 ××支店",
    "振込種別": "普通",
    "口座番号": "1234567",
    "口座名義": "株式会社SNSソフト"
  }
}
    '''
    json_data_3 = '''
    {
  "請求書": {
    "請求日": "2024/07/31",
    "請求番号": "226",
    "請求月": "2024年07月",
    "請求先": "東海ソフト株式会社 御中",
    "請求元": {
      "会社名": "株式会社SNSソフト",
      "住所": "〒101-0031 東京都千代田区東神田11番14 TQ東神田ビル2階",
      "電話番号": "03-6240-9295",
      "担当者": "尾崎 友理",
      "登録番号": "T6010001154383"
    },
    "件名": "システム設計・開発",
    "期間": "2024/07/01 ～ 2024/07/31",
    "支払い条件": "2024年8月31日",
    "合計金額": "1,626,592 円（税込）",
    "明細": [
      {
        "項目": "作業代 (張　以琳)",
        "数量": "1",
        "単位": "人月",
        "単価": "670,000",
        "金額": "670,000",
        "消費税率": "10%"
      },
      {
        "項目": "超過単価(190h)",
        "数量": "11",
        "単位": "時間",
        "単価": "3,520",
        "金額": "38,720",
        "消費税率": "10%"
      },
      {
        "項目": "控除単価(140h)",
        "数量": "0",
        "単位": "時間",
        "単価": "-4,780",
        "金額": "0",
        "消費税率": "10%"
      },
      {
        "項目": "作業代 (祝　勇)",
        "数量": "1",
        "単位": "人月",
        "単価": "770,000",
        "金額": "770,000",
        "消費税率": "10%"
      },
      {
        "項目": "超過単価(190h)",
        "数量": "0",
        "単位": "時間",
        "単価": "4,050",
        "金額": "0",
        "消費税率": "10%"
      },
      {
        "項目": "控除単価(140h)",
        "数量": "0",
        "単位": "時間",
        "単価": "-5,500",
        "金額": "0",
        "消費税率": "10%"
      }
    ],
    "小計": "1,478,720",
    "消費税": {
      "10%": "147,872",
      "8%": "0"
    },
    "合計": "1,626,592",
    "振込先": [
      "GMOあおぞらネット銀行法人第二営業部 普通 1284635 カ）エスエヌエスソフト",
      "みずほ銀行東京中央支店 普通 2433536 カ）エスエヌエスソフト"
    ],
    "備考": ""
  }
}
    '''
    
    ocr_data = '''
    [[[355.0, 27.0, 435.0, 48.0], '請求書'], [[604.0, 68.0, 724.0, 85.0], '講求日：2024/07/31'], [[163.0, 70.0, 191.0, 85.0], '226'], [[79.0, 72.0, 100.0, 85.0], 'no.'], [[164.0, 85.0, 234.0, 100.0], '2024年07月'], [[79.0, 86.0, 116.0, 100.0], '請求月'], [[68.0, 116.0, 254.0, 133.0], '東海ソフト株式会社  御中'], [[508.0, 116.0, 626.0, 133.0], '株式会社SNSソフト'], [[69.0, 145.0, 270.0, 163.0], '下記の通り、ご請求申し上げます。'], [[509.0, 149.0, 580.0, 163.0], '〒101-0031'], [[508.0, 164.0, 637.0, 180.0], '東京都千代田区東神由'], [[662.0, 165.0, 718.0, 179.0], '日11番14'], [[79.0, 175.0, 106.0, 190.0], '件名'], [[154.0, 175.0, 271.0, 192.0], 'システム設計・開発'], [[509.0, 180.0, 527.0, 193.0], 'fr'], [[79.0, 191.0, 105.0, 207.0], '期間'], [[153.0, 191.0, 322.0, 208.0], '2024/07/01 ～ 2024/07/31'], [[508.0, 194.0, 627.0, 208.0], 'TQ東神田ビル2階階'], [[80.0, 209.0, 134.0, 223.0], '支払い条件'], [[153.0, 209.0, 244.0, 224.0], '2024年8月31日'], [[508.0, 209.0, 627.0, 224.0], 'TEL：03-6240-9295'], [[508.0, 224.0, 618.0, 239.0], '担当： 尾崎  友理'], [[508.0, 239.0, 676.0, 254.0], '登録番号：T6010001154383'], [[80.0, 243.0, 135.0, 260.0], '合計金額'], [[162.0, 243.0, 284.0, 259.0], '1,626,592 円（税込）'], [[582.0, 394.0, 639.0, 409.0], '非課\u3000消費'], [[224.0, 400.0, 250.0, 414.0], '摘要'], [[470.0, 400.0, 498.0, 414.0], '実績'], [[527.0, 400.0, 553.0, 414.0], '単価'], [[669.0, 400.0, 695.0, 414.0], '金額'], [[406.0, 401.0, 466.0, 414.0], '数量 '], [[588.0, 408.0, 604.0, 422.0], '段'], [[619.0, 408.0, 637.0, 421.0], '&'], [[538.0, 419.0, 581.0, 436.0], '670,000'], [[71.0, 420.0, 157.0, 436.0], '作業代 (張\u3000以琳)'], [[438.0, 421.0, 468.0, 434.0], '人月'], [[480.0, 421.0, 502.0, 435.0], '201'], [[619.0, 421.0, 645.0, 434.0], '10%'], [[680.0, 421.0, 722.0, 434.0], '670,000'], [[426.0, 422.0, 439.0, 434.0], '1'], [[70.0, 434.0, 147.0, 448.0], '超過单価(190h)'], [[441.0, 434.0, 469.0, 448.0], '時間'], [[685.0, 434.0, 723.0, 448.0], '38,720'], [[619.0, 435.0, 645.0, 448.0], '10%'], [[421.0, 436.0, 438.0, 448.0], '11'], [[549.0, 436.0, 580.0, 448.0], '3,520'], [[71.0, 449.0, 147.0, 463.0], '控除単価(140h)'], [[442.0, 449.0, 468.0, 463.0], '時間'], [[548.0, 449.0, 580.0, 463.0], '-4,780'], [[619.0, 449.0, 644.0, 463.0], '10%'], [[711.0, 450.0, 723.0, 463.0], '0'], [[427.0, 451.0, 437.0, 462.0], '0'], [[680.0, 463.0, 723.0, 478.0], '770,000'], [[71.0, 464.0, 147.0, 478.0], '作業代 (祝\u3000勇)'], [[442.0, 464.0, 501.0, 478.0], '人月 183.75'], [[537.0, 464.0, 580.0, 478.0], '770,000'], [[619.0, 464.0, 644.0, 478.0], '10%'], [[427.0, 466.0, 436.0, 477.0], '1'], [[70.0, 478.0, 146.0, 493.0], '超過单価(190h)'], [[443.0, 479.0, 469.0, 492.0], '時間'], [[711.0, 479.0, 723.0, 493.0], '0'], [[427.0, 480.0, 438.0, 491.0], '0'], [[549.0, 480.0, 580.0, 492.0], '4,050'], [[619.0, 480.0, 645.0, 492.0], '10%'], [[71.0, 491.0, 146.0, 507.0], '控除単価(140h)'], [[442.0, 491.0, 469.0, 506.0], '時間'], [[426.0, 493.0, 437.0, 503.0], '0'], [[548.0, 492.0, 581.0, 505.0], '-5,500'], [[619.0, 492.0, 645.0, 506.0], '10%'], [[710.0, 492.0, 723.0, 506.0], '0'], [[384.0, 611.0, 416.0, 627.0], '小計'], [[488.0, 611.0, 626.0, 627.0], '1,478,720   消費税10%'], [[670.0, 611.0, 723.0, 627.0], '147,872'], [[385.0, 628.0, 429.0, 644.0], '消費税'], [[498.0, 628.0, 557.0, 644.0], '147,872'], [[555.0, 628.0, 619.0, 644.0], '消費税8%'], [[709.0, 628.0, 723.0, 645.0], '0'], [[386.0, 646.0, 416.0, 661.0], '合計'], [[488.0, 646.0, 551.0, 661.0], '1,626,592'], [[382.0, 716.0, 418.0, 729.0], '振込元'], [[80.0, 731.0, 477.0, 746.0], 'GMOあおぞらネット銀行法人第二営業部\u3000普通\u30001284635\u3000カ）エスエヌエスソフト'], [[80.0, 744.0, 403.0, 758.0], 'みずほ銀行東京中央支店\u3000普通\u30002433536\u3000カ）エスエヌエスソフト'], [[387.0, 767.0, 412.0, 781.0], '備考']]
    '''
    ocr_data = '''
    [[[401.0, 11.0, 491.0, 40.0], '請求書'], [[87.0, 92.0, 336.0, 116.0], '株式会社 00000'], [[355.0, 93.0, 423.0, 116.0], '御中'], [[772.0, 93.0, 869.0, 115.0], 'ABC0022'], [[584.0, 94.0, 690.0, 117.0], '請求書No.'], [[746.0, 125.0, 868.0, 147.0], '2023/11/30'], [[618.0, 126.0, 689.0, 149.0], '請求日'], [[26.0, 186.0, 633.0, 209.0], 'ご利用ありがとうございます。下記の通り御請求申し上げます。'], [[227.0, 244.0, 327.0, 268.0], '131,200'], [[106.0, 245.0, 179.0, 268.0], '11月分'], [[370.0, 245.0, 472.0, 268.0], '町（税込）'], [[277.0, 282.0, 324.0, 306.0], '品名'], [[80.0, 284.0, 126.0, 306.0], '日付'], [[535.0, 284.0, 582.0, 306.0], '単価'], [[627.0, 284.0, 674.0, 306.0], '数量'], [[708.0, 284.0, 754.0, 305.0], '単位'], [[798.0, 284.0, 845.0, 307.0], '金額'], [[436.0, 285.0, 484.0, 306.0], '税率'], [[59.0, 319.0, 141.0, 341.0], '11月1日'], [[804.0, 319.0, 863.0, 342.0], '5,000'], [[182.0, 320.0, 250.0, 341.0], '小麦粉'], [[384.0, 320.0, 409.0, 342.0], '※'], [[547.0, 320.0, 606.0, 342.0], '1,000'], [[643.0, 320.0, 660.0, 342.0], '5'], [[443.0, 321.0, 478.0, 342.0], '8%'], [[718.0, 322.0, 744.0, 342.0], '回'], [[59.0, 354.0, 141.0, 377.0], '11月1日'], [[182.0, 354.0, 230.0, 375.0], '牛肉'], [[547.0, 354.0, 606.0, 378.0], '2,500'], [[793.0, 354.0, 863.0, 377.0], '10,000'], [[384.0, 355.0, 409.0, 376.0], '※'], [[443.0, 355.0, 478.0, 376.0], '8%'], [[643.0, 355.0, 660.0, 376.0], '4'], [[719.0, 356.0, 744.0, 376.0], '個'], [[58.0, 389.0, 141.0, 412.0], '11月2日'], [[547.0, 389.0, 606.0, 412.0], '1,000'], [[804.0, 389.0, 863.0, 412.0], '2,000'], [[438.0, 390.0, 483.0, 411.0], '10%'], [[643.0, 390.0, 661.0, 412.0], '2'], [[183.0, 391.0, 352.0, 408.0], 'キッチンペーパー'], [[719.0, 391.0, 743.0, 411.0], '回'], [[182.0, 422.0, 356.0, 446.0], '業務用IH調理器具'], [[59.0, 424.0, 141.0, 446.0], '11月3日'], [[438.0, 424.0, 483.0, 447.0], '10%'], [[535.0, 424.0, 607.0, 446.0], '78,000'], [[719.0, 424.0, 745.0, 447.0], '式'], [[793.0, 424.0, 863.0, 447.0], '78,000'], [[643.0, 425.0, 660.0, 448.0], '1'], [[59.0, 459.0, 141.0, 481.0], '11月3日'], [[181.0, 459.0, 354.0, 480.0], '伊勢海老鍋セット'], [[442.0, 459.0, 478.0, 481.0], '8%'], [[547.0, 459.0, 607.0, 482.0], '5,000'], [[643.0, 459.0, 661.0, 481.0], '5'], [[792.0, 459.0, 863.0, 482.0], '25,000'], [[384.0, 460.0, 409.0, 480.0], '次'], [[696.0, 461.0, 762.0, 481.0], 'セット'], [[263.0, 703.0, 398.0, 727.0], '120,000 | 円'], [[679.0, 702.0, 799.0, 727.0], '11,200 円'], [[73.0, 704.0, 182.0, 727.0], '合計(税抜)'], [[460.0, 704.0, 572.0, 727.0], '消費税額計'], [[710.0, 740.0, 799.0, 763.0], '8,000円'], [[132.0, 741.0, 229.0, 764.0], '10% 対象'], [[298.0, 741.0, 398.0, 764.0], '80,000円'], [[544.0, 742.0, 611.0, 764.0], '消費税'], [[711.0, 775.0, 799.0, 798.0], '3,200円'], [[543.0, 776.0, 612.0, 799.0], '消費税'], [[144.0, 777.0, 229.0, 798.0], '8% 対象'], [[299.0, 777.0, 398.0, 798.0], '40,000円'], [[351.0, 808.0, 399.0, 836.0], 'の円'], [[752.0, 810.0, 799.0, 834.0], '0 円'], [[144.0, 811.0, 229.0, 833.0], '0% 対象'], [[543.0, 811.0, 612.0, 834.0], '消費税'], [[24.0, 848.0, 176.0, 868.0], '※軽減税率対象'], [[551.0, 879.0, 867.0, 902.0], '△△商事株式会社・購買部（XX）'], [[531.0, 914.0, 868.0, 937.0], '〒10X-XXX\u3000東京都港区XX1-2-4'], [[657.0, 950.0, 868.0, 971.0], 'TEL：03-1234-XXXX'], [[667.0, 984.0, 856.0, 1006.0], 'T 1234567890XXX'], [[522.0, 985.0, 613.0, 1008.0], '登録番号']]
    '''

    # json_to_excel(json_data_2, 'output_2.xlsx')
    # json_to_excel_with_links(json_data_2, 'output_with_links.xlsx')
    json_result = get_json_result(ocr_data)
    print(json_result)
    json_to_excel_with_links(json_result, 'output_with_links_2.xlsx')
